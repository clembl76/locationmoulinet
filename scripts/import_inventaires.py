"""
Import des inventaires Excel → Supabase
Tables: apartment_installation, apartment_keys, check_in_elements, inventory_items
"""

import openpyxl
import re
import json
import requests
from pathlib import Path

# ─── Config ───────────────────────────────────────────────────────────────────

EXCEL_PATH = Path('C:/Users/cleme/gestion-locative/data/InventairesActuels.xlsx')
ENV_PATH   = Path('C:/Users/cleme/locationmoulinet/.env.local')

env = {}
for line in ENV_PATH.read_text(encoding='utf-8').splitlines():
    if '=' in line and not line.startswith('#'):
        k, v = line.split('=', 1)
        env[k.strip()] = v.strip().strip('"').strip("'")

SUPABASE_URL = env['NEXT_PUBLIC_SUPABASE_URL']
SUPABASE_KEY = env['SUPABASE_SERVICE_ROLE_KEY']

HEADERS = {
    'apikey': SUPABASE_KEY,
    'Authorization': f'Bearer {SUPABASE_KEY}',
    'Content-Type': 'application/json',
    'Prefer': 'return=minimal',
}

# ─── Normalisation ────────────────────────────────────────────────────────────

VALID_PIECES = {
    'Chambre', 'Coin chambre', 'Coin cuisine', 'Coin nuit', 'Coin salon',
    'Coin salon / salle à manger', 'Couloir', 'Cuisine', 'Divers', 'Entrée',
    'Parties communes', 'Partout', 'Salle à manger', 'Salle de bains',
    'Salle de douche', 'Salon', 'Séjour/Cuisine', 'Terrasse', 'Toilettes',
}

ETAT_MAP = {
    "Neuf":                         "Neuf",
    "Bon état":                     "Bon état",
    "Etat d'usage":                 "État d'usage",
    "État d'usage":                 "État d'usage",
    "Mauvais état":                 "Mauvais état",
    "Etat d'usage/Peinture Neuve":  "État d'usage",
}

SKIP_INVENTORY_PIECES = {
    "Pièce", "INVENTAIRE", "ETAT DES LIEUX", "Adresse", "Appartement",
    "Date EDL Entrée", "Date EDL Sortie", "Entrée dans les lieux le :",
    "Sortie des lieux le :", "Mode :", "Montant de la caution versée :",
    "Fait en 2 exemplaires, à Rouen le", "Eau chaude", "Chauffage",
    "Eau", "Electricité",
}

def s(v):
    """Convertit une valeur Excel en str propre, None si vide."""
    if v is None:
        return None
    t = str(v).strip()
    return t if t else None

def norm_etat(v):
    raw = s(v)
    if not raw:
        return None
    return ETAT_MAP.get(raw)  # None si valeur inconnue

def norm_piece(v):
    raw = s(v)
    if not raw or raw in SKIP_INVENTORY_PIECES:
        return None
    return raw if raw in VALID_PIECES else None

# ─── Supabase helpers ─────────────────────────────────────────────────────────

def sb_get(table, params=''):
    r = requests.get(f'{SUPABASE_URL}/rest/v1/{table}?{params}', headers=HEADERS)
    r.raise_for_status()
    return r.json()

def sb_insert(table, rows, chunk=100):
    """Insert par lots de 100 pour éviter les timeouts."""
    inserted = 0
    for i in range(0, len(rows), chunk):
        batch = rows[i:i+chunk]
        r = requests.post(
            f'{SUPABASE_URL}/rest/v1/{table}',
            headers={**HEADERS, 'Prefer': 'return=minimal'},
            json=batch,
        )
        if r.status_code not in (200, 201):
            print(f'  ✗ Erreur {table}: {r.status_code} — {r.text[:200]}')
        else:
            inserted += len(batch)
    return inserted

def sb_delete(table, apt_id):
    r = requests.delete(
        f'{SUPABASE_URL}/rest/v1/{table}?apartment_id=eq.{apt_id}',
        headers=HEADERS,
    )
    return r.status_code

# ─── Parser ───────────────────────────────────────────────────────────────────

def parse_sheet(ws):
    rows = list(ws.iter_rows(values_only=True))

    # Trouver les index des sections
    edl_header_idx = None    # ligne 'Pièce / Element'
    inv_header_idx = None    # ligne 'Pièce / Article'
    keys_start_idx = None    # ligne 'Eau chaude'

    for i, row in enumerate(rows):
        r0 = s(row[0]) or ''
        r1 = s(row[1]) or ''
        if r0 == 'Eau chaude':
            keys_start_idx = i
        if r0 == 'Pièce' and r1 == 'Element':
            edl_header_idx = i
        if r0 == 'Pièce' and r1 == 'Article':
            inv_header_idx = i

    # ── Installation + clés ──
    installation = {}
    keys = []

    if keys_start_idx is not None:
        end = edl_header_idx if edl_header_idx else len(rows)
        for i in range(keys_start_idx, end):
            row = rows[i]
            r0, r1, r4, r5 = s(row[0]), s(row[1]), s(row[4]), row[5]

            if r0 == 'Eau chaude':
                installation['hot_water'] = r1
            elif r0 == 'Chauffage':
                installation['heating'] = r1

            # Clés : col E (index 4) = type, col F (index 5) = quantité
            if r4 and r4 not in ('Inventaire et remise des clefs',):
                try:
                    qty = int(float(r5)) if r5 is not None else 0
                except (ValueError, TypeError):
                    qty = 0
                keys.append({'key_type': r4, 'quantity': qty, 'order_index': len(keys)})

    # ── EDL éléments ──
    edl_items = []
    if edl_header_idx is not None:
        inv_start = inv_header_idx if inv_header_idx else len(rows)
        current_piece = None
        for i in range(edl_header_idx + 1, inv_start):
            row = rows[i]
            r0 = s(row[0])
            r1 = s(row[1])

            # Fin de section
            if r0 == 'INVENTAIRE':
                break

            # Mise à jour de la pièce courante
            piece = norm_piece(r0)
            if piece:
                current_piece = piece

            if not r1 or not current_piece:
                continue
            if r1 in ('Element', 'Entrée'):
                continue

            edl_items.append({
                'room':            current_piece,
                'element':         r1,
                'condition_entry': norm_etat(row[2]),
                'comment_entry':   s(row[3]),
                'condition_exit':  norm_etat(row[4]),
                'comment_exit':    s(row[5]),
                'order_index':     len(edl_items),
            })

    # ── Inventaire articles ──
    inv_items = []
    if inv_header_idx is not None:
        current_piece = None
        for i in range(inv_header_idx + 1, len(rows)):
            row = rows[i]
            r0 = s(row[0])
            r1 = s(row[1])

            if not r1:
                continue
            if r1 in ('Article', 'Quantité'):
                continue

            # Pièce
            piece = norm_piece(r0)
            if piece:
                current_piece = piece
            elif r0 and r0 not in SKIP_INVENTORY_PIECES:
                # Pièce inconnue → skip cette ligne
                pass

            if not current_piece:
                continue

            # Quantités
            try:
                qty_in = int(float(row[2])) if row[2] is not None else None
            except (ValueError, TypeError):
                qty_in = None
            try:
                qty_out = int(float(row[4])) if row[4] is not None else None
            except (ValueError, TypeError):
                qty_out = None

            inv_items.append({
                'room':            current_piece,
                'item':            r1.strip(),
                'quantity_entry':  qty_in,
                'condition_entry': norm_etat(row[3]),
                'comment_entry':   s(row[3]) if norm_etat(row[3]) is None and s(row[3]) else None,
                'quantity_exit':   qty_out,
                'condition_exit':  norm_etat(row[5]),
                'comment_exit':    s(row[5]) if norm_etat(row[5]) is None and s(row[5]) else None,
                'order_index':     len(inv_items),
            })

    return installation, keys, edl_items, inv_items

# ─── Main ─────────────────────────────────────────────────────────────────────

def main():
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    apt_sheets = [s for s in wb.sheetnames if re.match(r'^\d+$', s)]

    # Charger la map number→id depuis Supabase
    apts = sb_get('apartments', 'select=id,number')
    apt_map = {a['number']: a['id'] for a in apts}
    print(f'Appartements Supabase : {sorted([k for k in apt_map.keys() if k.isdigit()], key=int)}')

    total_edl = total_inv = total_keys = total_install = 0

    for sheet_name in apt_sheets:
        apt_id = apt_map.get(sheet_name)
        if not apt_id:
            print(f'\n⚠ Onglet {sheet_name} : pas d\'appartement correspondant dans Supabase, ignoré.')
            continue

        ws = wb[sheet_name]
        installation, keys, edl_items, inv_items = parse_sheet(ws)

        print(f'\nAppt {sheet_name} (id={apt_id[:8]}…)')
        print(f'  Installation: {installation}')
        print(f'  Clés: {len(keys)} types')
        print(f'  EDL éléments: {len(edl_items)}')
        print(f'  Inventaire: {len(inv_items)}')

        # Supprimer les données existantes (idempotent)
        for table in ['apartment_installation', 'apartment_keys', 'check_in_elements', 'inventory_items']:
            sb_delete(table, apt_id)

        # Insérer installation
        if installation:
            sb_insert('apartment_installation', [{'apartment_id': apt_id, **installation}])
            total_install += 1

        # Insérer clés
        if keys:
            for k in keys:
                k['apartment_id'] = apt_id
            sb_insert('apartment_keys', keys)
            total_keys += len(keys)

        # Insérer EDL éléments
        if edl_items:
            for item in edl_items:
                item['apartment_id'] = apt_id
            sb_insert('check_in_elements', edl_items)
            total_edl += len(edl_items)

        # Insérer inventaire
        if inv_items:
            for item in inv_items:
                item['apartment_id'] = apt_id
            sb_insert('inventory_items', inv_items)
            total_inv += len(inv_items)

    print(f'\nImport termine')
    print(f'  Installations : {total_install}')
    print(f'  Cles          : {total_keys}')
    print(f'  EDL elements  : {total_edl}')
    print(f'  Inventaire    : {total_inv}')

if __name__ == '__main__':
    main()
